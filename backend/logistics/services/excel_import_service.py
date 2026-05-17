from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth import get_user_model

from logistics.models import Aliado, Cliente, Pedido, Repartidor, Ruta, RutaParada


SHEET_ALIASES = {
    'aliados': 'aliados',
    'bodegas': 'aliados',
    'tiendas': 'aliados',
    'repartidores': 'repartidores',
    'drivers': 'repartidores',
    'clientes': 'clientes',
    'pedidos': 'pedidos',
    'datos_rutas': 'pedidos',
    'datos_ruta': 'pedidos',
    'rutas': 'pedidos',
    'ruta': 'pedidos',
}


def import_excel_file(file_path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('Falta instalar openpyxl. Ejecuta: pip install openpyxl') from exc

    path = Path(file_path)
    if path.suffix.lower() not in {'.xlsx', '.xls'}:
        return _result(errors=[f'Extension no permitida: {path.suffix}. Use .xlsx o .xls.'])

    result = _result()
    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        result['errors'].append(f'No se pudo leer el Excel: {exc}')
        return result

    seen_sheets = set()
    imported_pedidos = []
    for sheet in workbook.worksheets:
        rows = list(_rows_as_dicts(sheet))
        if not rows:
            continue

        sheet_type = _sheet_type(sheet.title, rows[0])
        if sheet_type not in SHEET_ALIASES.values():
            result['warnings'].append(f'Hoja ignorada: {sheet.title}')
            continue

        seen_sheets.add(sheet_type)
        for index, row in enumerate(rows, start=2):
            try:
                if sheet_type == 'aliados':
                    _import_aliado(row, result)
                elif sheet_type == 'repartidores':
                    _import_repartidor(row, result)
                elif sheet_type == 'clientes':
                    _import_cliente(row, result)
                elif sheet_type == 'pedidos':
                    p = _import_pedido(row, result)
                    if p:
                        imported_pedidos.append(p)
            except Exception as exc:
                result['errors'].append(f'Hoja {sheet.title}, fila {index}: {exc}')

    if imported_pedidos:
        _generate_routes_for_imported_pedidos(imported_pedidos, result)

    for expected in ('aliados', 'repartidores', 'clientes', 'pedidos'):
        if expected not in seen_sheets:
            result['warnings'].append(f'No se encontro hoja de {expected}; se continuo con las disponibles.')

    return result


def _result(errors=None):
    return {
        'message': 'Importacion completada',
        'created': {'aliados': 0, 'repartidores': 0, 'clientes': 0, 'pedidos': 0, 'rutas': 0},
        'updated': {'aliados': 0, 'repartidores': 0, 'clientes': 0},
        'errors': errors or [],
        'warnings': [],
    }


def _rows_as_dicts(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = [_normalize(header) for header in rows[0]]
    for row in rows[1:]:
        data = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(value not in (None, '') for value in data.values()):
            yield data


def _sheet_type(title, first_row):
    normalized_title = _normalize(title)
    if normalized_title in SHEET_ALIASES:
        return SHEET_ALIASES[normalized_title]
    row_type = _normalize(first_row.get('tipo') or first_row.get('type'))
    return SHEET_ALIASES.get(row_type, normalized_title)


def _import_aliado(row, result):
    nombre = _value(row, 'nombre', 'name', default='Bodega sin nombre')
    username = _value(row, 'username', 'usuario', default=_username(nombre, 'aliado'))
    user, user_created = _upsert_user(username, nombre, 'aliado')
    _, created = Aliado.objects.update_or_create(
        user=user,
        defaults={
            'direccion': _value(row, 'direccion', 'address', default='Sin direccion'),
            'latitud': _decimal(row, 'latitud', 'latitude'),
            'longitud': _decimal(row, 'longitud', 'longitude'),
        },
    )
    result['created' if created or user_created else 'updated']['aliados'] += 1


def _import_repartidor(row, result):
    nombre = _value(row, 'nombre', 'name', default='Repartidor sin nombre')
    username = _value(row, 'username', 'usuario', default=_username(nombre, 'driver'))
    user, user_created = _upsert_user(username, nombre, 'driver', estado=_value(row, 'estado', 'status', default='Disponible'))
    _, created = Repartidor.objects.update_or_create(
        user=user,
        defaults={
            'telefono': _value(row, 'telefono', 'phone', default=''),
            'latitud_actual': _decimal(row, 'latitud_actual', 'latitud', 'latitude'),
            'longitud_actual': _decimal(row, 'longitud_actual', 'longitud', 'longitude'),
            'capacidad_maxima_kg': _decimal(row, 'capacidad_maxima_kg', 'capacidad', default=15),
        },
    )
    result['created' if created or user_created else 'updated']['repartidores'] += 1


def _import_cliente(row, result):
    _, created = _upsert_cliente(row)
    result['created' if created else 'updated']['clientes'] += 1


def _clean_prioridad(val):
    if not val:
        return 'normal'
    val_lower = str(val).strip().lower()
    if val_lower in ('baja', 'low'):
        return 'baja'
    if val_lower in ('media', 'normal', 'medium', 'medio'):
        return 'normal'
    if val_lower in ('alta', 'high'):
        return 'alta'
    if val_lower in ('urgente', 'critical', 'urgent'):
        return 'urgente'
    return 'normal'


def _find_repartidor(row):
    repartidor_name = _value(row, 'repartidor', 'driver', 'conductor', 'repartidor_username', 'driver_username', default=None)
    if not repartidor_name:
        return None
    
    repartidor_name = str(repartidor_name).strip()
    User = get_user_model()
    # Search by username (exact)
    user = User.objects.filter(role='driver', username__iexact=repartidor_name).first()
    if not user:
        # Search by nombre (contains)
        user = User.objects.filter(role='driver', nombre__icontains=repartidor_name).first()
    return user


def _import_pedido(row, result):
    cliente, client_created = _upsert_cliente(row)
    if client_created:
        result['created']['clientes'] += 1
    
    repartidor = _find_repartidor(row)
    estado = _value(row, 'estado', 'status', default='Pendiente')
    if repartidor and estado == 'Pendiente':
        estado = 'Asignado'
        
    pedido = Pedido.objects.create(
        cliente=cliente,
        aliado=_find_aliado(row),
        repartidor=repartidor,
        descripcion=_value(row, 'descripcion', 'description', default=''),
        estado=estado,
        prioridad=_clean_prioridad(_value(row, 'prioridad', 'priority', default='normal')),
        peso_total_kg=_decimal(row, 'peso_total_kg', 'peso', 'weightkg', default=0),
        volumen_total_m3=_decimal(row, 'volumen_total_m3', 'volumen', default=None),
    )
    result['created']['pedidos'] += 1
    return pedido


def haversine_km(lat1, lon1, lat2, lon2):
    radius_km = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return radius_km * c


def _generate_routes_for_imported_pedidos(pedidos, result):
    from django.db import transaction
    from collections import defaultdict
    from decimal import Decimal
    from logistics.models import Ruta, RutaParada, Repartidor

    # Group orders that have a driver assigned
    by_driver = defaultdict(list)
    for p in pedidos:
        if p.repartidor_id:
            by_driver[p.repartidor].append(p)

    if not by_driver:
        return

    result['created']['rutas'] = 0

    with transaction.atomic():
        for driver, driver_pedidos in by_driver.items():
            # Get driver start location
            profile = Repartidor.objects.filter(user=driver).first()
            if profile and profile.latitud_actual is not None and profile.longitud_actual is not None:
                start_lat = float(profile.latitud_actual)
                start_lng = float(profile.longitud_actual)
            else:
                # Use centroid or first order
                first_order = driver_pedidos[0]
                start_lat = float(first_order.cliente.latitud or 4.7110)
                start_lng = float(first_order.cliente.longitud or -74.0721)

            # Sort orders using nearest neighbor
            remaining = list(driver_pedidos)
            current_lat = start_lat
            current_lng = start_lng
            ordered_stops = []
            total_distance = 0.0

            while remaining:
                # Find nearest order
                next_order = min(
                    remaining,
                    key=lambda p: haversine_km(
                        current_lat,
                        current_lng,
                        float(p.cliente.latitud or current_lat),
                        float(p.cliente.longitud or current_lng),
                    )
                )
                next_lat = float(next_order.cliente.latitud or current_lat)
                next_lng = float(next_order.cliente.longitud or current_lng)
                leg_distance = haversine_km(current_lat, current_lng, next_lat, next_lng)
                total_distance += leg_distance

                ordered_stops.append({
                    'pedido': next_order,
                    'lat': next_lat,
                    'lng': next_lng,
                    'distance': leg_distance,
                })
                current_lat = next_lat
                current_lng = next_lng
                remaining.remove(next_order)

            # Estimate duration (average speed 28 km/h, 4 mins service time per stop)
            travel_mins = (total_distance / 28.0) * 60
            service_mins = len(ordered_stops) * 4
            total_duration_mins = int(math.ceil(travel_mins + service_mins))

            # Build geometry
            geometry = {
                'type': 'LineString',
                'coordinates': [[start_lng, start_lat]] + [
                    [stop['lng'], stop['lat']] for stop in ordered_stops
                ],
            }

            # Create Route
            aliado = driver_pedidos[0].aliado
            route = Ruta.objects.create(
                repartidor=driver,
                aliado=aliado,
                latitud_inicio=Decimal(str(round(start_lat, 8))),
                longitud_inicio=Decimal(str(round(start_lng, 8))),
                tiempo_estimado_mins=total_duration_mins,
                distancia_km=Decimal(str(round(total_distance, 2))),
                capacidad_usada_kg=Decimal(str(round(sum(float(p.peso_total_kg or 0) for p in driver_pedidos), 2))),
                estado_ruta='asignada',
                geometria=geometry,
                decision_ai={
                    'explicacion': f'Ruta de Excel generada automáticamente para {driver.nombre}. Total {len(driver_pedidos)} pedidos.',
                    'metrics': {
                        'distancia_total_km': round(total_distance, 2),
                        'tiempo_estimado_mins': total_duration_mins,
                        'eficiencia': 'Alta'
                    }
                }
            )

            # Create stops
            for index, stop in enumerate(ordered_stops, start=1):
                p = stop['pedido']
                # Calculate duration for this leg
                leg_travel_mins = (stop['distance'] / 28.0) * 60
                leg_duration = int(math.ceil(leg_travel_mins))

                RutaParada.objects.create(
                    ruta=route,
                    pedido=p,
                    orden=index,
                    latitud=Decimal(str(round(stop['lat'], 8))),
                    longitud=Decimal(str(round(stop['lng'], 8))),
                    distancia_desde_anterior_km=Decimal(str(round(stop['distance'], 2))),
                    tiempo_estimado_desde_anterior_mins=leg_duration,
                    estado='pendiente'
                )

                # Set order state to 'Asignado'
                p.repartidor = driver
                p.estado = 'Asignado'
                p.save(update_fields=['repartidor', 'estado'])

            # Set driver state to 'Ocupado'
            driver.estado = 'Ocupado'
            driver.save(update_fields=['estado'])

            result['created']['rutas'] += 1


def _upsert_user(username, nombre, role, estado='Disponible'):
    User = get_user_model()
    user, created = User.objects.get_or_create(
        username=username,
        defaults={'nombre': nombre, 'role': role, 'estado': estado},
    )
    user.nombre = nombre or user.nombre
    user.role = role
    user.estado = estado or user.estado
    if created:
        user.set_password('123')
    user.save()
    return user, created


def _upsert_cliente(row):
    nombre = _value(row, 'cliente', 'cliente_nombre', 'nombre', 'name', default='Cliente sin nombre')
    return Cliente.objects.update_or_create(
        nombre=nombre,
        defaults={
            'correo': _value(row, 'correo', 'email', default=None),
            'telefono': _value(row, 'telefono', 'phone', default=None),
            'direccion': _value(row, 'direccion', 'destination', 'address', default='Sin direccion'),
            'latitud': _decimal(row, 'latitud', 'latitude'),
            'longitud': _decimal(row, 'longitud', 'longitude'),
        },
    )


def _find_aliado(row):
    aliado_name = _value(row, 'aliado', 'bodega', 'tienda', default=None)
    if not aliado_name:
        return None
    return Aliado.objects.select_related('user').filter(user__nombre=aliado_name).first()


def _value(row, *keys, default=None):
    for key in keys:
        value = row.get(_normalize(key))
        if value not in (None, ''):
            return str(value).strip()
    return default


def _decimal(row, *keys, default=None):
    raw = _value(row, *keys, default=default)
    if raw in (None, ''):
        return None
    try:
        return Decimal(str(raw).replace(',', '.'))
    except (InvalidOperation, ValueError):
        return default


def _username(value, prefix):
    base = ''.join(char.lower() if char.isalnum() else '_' for char in value).strip('_')
    return f'{prefix}_{base[:40] or "importado"}'


def _normalize(value):
    if value is None:
        return ''
    return str(value).strip().lower().replace(' ', '_')
