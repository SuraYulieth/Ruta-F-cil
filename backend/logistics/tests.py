from django.test import TestCase
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APIClient
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import skipIf

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

from .models import Aliado, Cliente, CustomUser, Pedido, Repartidor, Ruta, RutaParada
from .services.route_metrics_service import RouteMetricsService
from .services.route_optimizer_service import RouteOptimizerService, haversine_km


class RouteOptimizerServiceTests(TestCase):
    def setUp(self):
        self.driver = CustomUser.objects.create_user(
            username='driver1',
            password='123',
            role='driver',
            nombre='Driver Uno',
        )
        Repartidor.objects.create(
            user=self.driver,
            latitud_actual=4.7100,
            longitud_actual=-74.0700,
            capacidad_maxima_kg=10,
            disponible=True,
        )
        warehouse_user = CustomUser.objects.create_user(
            username='warehouse1',
            password='123',
            role='aliado',
            nombre='Bodega Test',
        )
        self.warehouse = Aliado.objects.create(
            user=warehouse_user,
            direccion='Bodega',
            latitud=4.7120,
            longitud=-74.0710,
        )
        cliente_a = Cliente.objects.create(
            nombre='Cliente A',
            direccion='Zona A',
            latitud=4.7110,
            longitud=-74.0721,
        )
        cliente_b = Cliente.objects.create(
            nombre='Cliente B',
            direccion='Zona B',
            latitud=4.7150,
            longitud=-74.0750,
        )
        Cliente.objects.create(
            nombre='Sin coordenadas',
            direccion='Sin datos',
        )
        self.pedido_a = Pedido.objects.create(cliente=cliente_a, prioridad='alta', peso_total_kg=2)
        self.pedido_b = Pedido.objects.create(cliente=cliente_b, prioridad='normal', peso_total_kg=3)

    def test_haversine_returns_positive_distance(self):
        distance = haversine_km(4.7110, -74.0721, 4.7150, -74.0750)
        self.assertGreater(distance, 0)

    def test_optimizer_selects_feasible_orders(self):
        result = RouteOptimizerService().optimize(
            repartidor_id=self.driver.id,
            latitud_inicial=4.7100,
            longitud_inicial=-74.0700,
            capacidad_maxima=10,
        )

        selected_ids = [pedido.id for pedido in result['pedidos_seleccionados']]
        self.assertIn(self.pedido_a.id, selected_ids)
        self.assertIn(self.pedido_b.id, selected_ids)
        self.assertEqual(len(result['orden_entrega']), 2)
        self.assertGreater(result['distancia_total_km'], 0)
        self.assertEqual(result['aliado_id'], self.warehouse.id)

    def test_optimizer_can_select_driver_automatically(self):
        result = RouteOptimizerService().optimize(capacidad_maxima=10)
        self.assertEqual(result['repartidor_id'], self.driver.id)

    def test_optimizer_multi_route_creates_multiple_routes_without_repeating_orders(self):
        second_driver = CustomUser.objects.create_user(
            username='driver2',
            password='123',
            role='driver',
            nombre='Driver Dos',
        )
        Repartidor.objects.create(
            user=second_driver,
            latitud_actual=4.8000,
            longitud_actual=-74.2000,
            capacidad_maxima_kg=10,
            disponible=True,
        )

        cliente_c = Cliente.objects.create(
            nombre='Cliente C',
            direccion='Zona C',
            latitud=4.8005,
            longitud=-74.1995,
        )
        pedido_c = Pedido.objects.create(cliente=cliente_c, prioridad='alta', peso_total_kg=2)

        response = self.client.post('/api/routes/optimize/', {
            'modo': 'multi_ruta',
            'pedidos_candidatos': [self.pedido_a.id, self.pedido_b.id, pedido_c.id],
            'max_duration_mins': 90,
            'max_area_km2': 382,
        }, format='json')

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['summary']['rutas_creadas'], 2)
        self.assertEqual(len(response.data['routes']), 2)

        all_selected = response.data['optimizer']['pedidos_seleccionados']
        self.assertCountEqual(all_selected, [self.pedido_a.id, self.pedido_b.id, pedido_c.id])
        self.assertEqual(len(all_selected), len(set(all_selected)))
        self.assertEqual(Ruta.objects.count(), 2)
        self.assertEqual(RutaParada.objects.count(), 3)


class RouteApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = CustomUser.objects.create_user(
            username='admin-tests',
            password='123',
            role='admin',
            nombre='Admin Tests',
        )
        self.client.force_authenticate(user=self.admin)
        self.driver = CustomUser.objects.create_user(
            username='driver-api',
            password='123',
            role='driver',
            nombre='Driver API',
        )
        Repartidor.objects.create(
            user=self.driver,
            latitud_actual=4.7100,
            longitud_actual=-74.0700,
            capacidad_maxima_kg=5,
            disponible=True,
        )
        warehouse_user = CustomUser.objects.create_user(
            username='warehouse-api',
            password='123',
            role='aliado',
            nombre='Bodega API',
        )
        Aliado.objects.create(
            user=warehouse_user,
            direccion='Centro',
            latitud=4.7115,
            longitud=-74.0715,
        )
        cliente = Cliente.objects.create(
            nombre='Cliente API',
            direccion='Centro',
            latitud=4.7110,
            longitud=-74.0721,
        )
        self.pedido = Pedido.objects.create(cliente=cliente, prioridad='urgente', peso_total_kg=1)

    def test_optimize_route_endpoint_creates_route(self):
        response = self.client.post('/api/routes/optimize/', {
            'repartidor_id': self.driver.id,
            'latitud_inicial': 4.7100,
            'longitud_inicial': -74.0700,
            'pedidos_candidatos': [self.pedido.id],
            'capacidad_maxima': 5,
        }, format='json')

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['optimizer']['pedidos_seleccionados'], [self.pedido.id])
        self.assertEqual(len(response.data['route']['paradas']), 1)
        self.assertIn('metrics', response.data)

    def test_optimize_route_endpoint_accepts_auto_driver(self):
        response = self.client.post('/api/routes/optimize/', {
            'repartidor_id': 'auto',
            'pedidos_candidatos': [self.pedido.id],
            'capacidad_maxima': 5,
        }, format='json')

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['optimizer']['pedidos_seleccionados'], [self.pedido.id])
        self.assertEqual(response.data['route']['repartidor'], self.driver.id)

    def test_optimize_route_endpoint_accepts_null_driver(self):
        response = self.client.post('/api/routes/optimize/', {
            'repartidor_id': None,
            'pedidos_candidatos': [self.pedido.id],
            'capacidad_maxima': 5,
        }, format='json')

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['optimizer']['pedidos_seleccionados'], [self.pedido.id])
        self.assertEqual(response.data['route']['repartidor'], self.driver.id)

    def test_optimize_multi_route_without_available_drivers_returns_unassigned_orders(self):
        Repartidor.objects.all().delete()

        response = self.client.post('/api/routes/optimize/', {
            'modo': 'multi_ruta',
            'pedidos_candidatos': [self.pedido.id],
            'capacidad_maxima': 5,
        }, format='json')

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.data['route'])
        self.assertEqual(response.data['routes'], [])
        self.assertEqual(response.data['summary']['rutas_creadas'], 0)
        self.assertEqual(response.data['summary']['pedidos_no_asignados'], 1)
        self.assertEqual(response.data['optimizer']['modo'], 'multi_ruta')
        self.assertEqual(response.data['optimizer']['pedidos_seleccionados'], [])
        self.assertEqual(len(response.data['optimizer']['unassigned_orders']), 1)
        self.assertEqual(response.data['optimizer']['unassigned_orders'][0]['pedido_id'], self.pedido.id)

    def test_manual_assign_endpoint_updates_status_and_driver(self):
        response = self.client.post(
            f'/api/pedidos/{self.pedido.id}/assign/',
            {'repartidor_id': self.driver.id},
            format='json',
        )

        self.assertEqual(response.status_code, 200)
        self.pedido.refresh_from_db()
        self.assertEqual(self.pedido.estado, 'Asignado')
        self.assertEqual(self.pedido.repartidor_id, self.driver.id)
        self.assertEqual(response.data['pedido']['estado'], 'Asignado')
        self.assertEqual(response.data['pedido']['id'], self.pedido.id)


class RouteMetricsServiceTests(TestCase):
    def test_build_returns_abp_evidence(self):
        result = {
            'pedidos_seleccionados': [],
            'distancia_total_km': 10,
            'duracion_total_mins': 30,
            'capacidad_usada_kg': 5,
            'aliado_nombre': 'Bodega',
            'repartidor_nombre': 'Driver',
            'explicacion': 'Decision explicada',
        }
        metrics = RouteMetricsService().build(result)
        self.assertEqual(metrics['complejidad']['caso_promedio'], 'O(n^2)')
        self.assertEqual(metrics['comparacion_manual_excel']['distancia_sistema_km'], 10)


@skipIf(Workbook is None, 'openpyxl no esta instalado')
class ImportExcelCommandTests(TestCase):
    def setUp(self):
        self.admin = CustomUser.objects.create_user(
            username='admin-import-tests',
            password='123',
            role='admin',
            nombre='Admin Import Tests',
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.admin)

    def test_import_excel_command_loads_core_entities(self):
        workbook = Workbook()
        aliados = workbook.active
        aliados.title = 'Aliados'
        aliados.append(['nombre', 'username', 'direccion', 'latitud', 'longitud'])
        aliados.append(['Bodega Uno', 'bodega_uno', 'Centro', 4.71, -74.07])

        repartidores = workbook.create_sheet('Repartidores')
        repartidores.append(['nombre', 'username', 'estado', 'latitud', 'longitud', 'capacidad'])
        repartidores.append(['Driver Uno', 'driver_uno', 'Disponible', 4.72, -74.08, 12])

        pedidos = workbook.create_sheet('Pedidos')
        pedidos.append(['cliente', 'direccion', 'latitud', 'longitud', 'prioridad', 'peso', 'bodega'])
        pedidos.append(['Cliente Uno', 'Calle 1', 4.73, -74.09, 'alta', 2, 'Bodega Uno'])

        with TemporaryDirectory() as directory:
            file_path = Path(directory) / 'import.xlsx'
            workbook.save(file_path)
            call_command('import_excel', str(file_path))

        self.assertEqual(Aliado.objects.count(), 1)
        self.assertEqual(Repartidor.objects.count(), 1)
        self.assertEqual(Cliente.objects.count(), 1)
        self.assertEqual(Pedido.objects.count(), 1)

    def test_import_excel_endpoint_accepts_multipart_file(self):
        workbook = Workbook()
        clientes = workbook.active
        clientes.title = 'Clientes'
        clientes.append(['nombre', 'direccion', 'latitud', 'longitud'])
        clientes.append(['Cliente API Excel', 'Calle 2', 4.71, -74.07])

        with TemporaryDirectory() as directory:
            file_path = Path(directory) / 'import.xlsx'
            workbook.save(file_path)
            upload = SimpleUploadedFile(
                'import.xlsx',
                file_path.read_bytes(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )

        response = self.client.post('/api/import-excel/', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created']['clientes'], 1)
        self.assertEqual(Cliente.objects.count(), 1)

    def test_import_excel_endpoint_requires_file(self):
        response = self.client.post('/api/import-excel/', {}, format='multipart')

        self.assertEqual(response.status_code, 400)


class DriverApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()

        self.driver = CustomUser.objects.create_user(
            username='driver-flow',
            password='123',
            role='driver',
            nombre='Driver Flow',
            estado='Disponible',
        )
        self.repartidor = Repartidor.objects.create(
            user=self.driver,
            latitud_actual=4.7100,
            longitud_actual=-74.0700,
            capacidad_maxima_kg=20,
            disponible=True,
        )

        self.other_driver = CustomUser.objects.create_user(
            username='driver-other',
            password='123',
            role='driver',
            nombre='Driver Other',
            estado='Disponible',
        )
        Repartidor.objects.create(
            user=self.other_driver,
            latitud_actual=4.7200,
            longitud_actual=-74.0800,
            capacidad_maxima_kg=15,
            disponible=True,
        )

        self.cliente_a = Cliente.objects.create(
            nombre='Cliente Uno',
            direccion='Calle 1',
            latitud=4.7110,
            longitud=-74.0720,
        )
        self.cliente_b = Cliente.objects.create(
            nombre='Cliente Dos',
            direccion='Calle 2',
            latitud=4.7130,
            longitud=-74.0730,
        )

        self.order_assigned = Pedido.objects.create(
            cliente=self.cliente_a,
            repartidor=self.driver,
            estado='Asignado',
            prioridad='alta',
            peso_total_kg=2,
        )
        self.order_other_driver = Pedido.objects.create(
            cliente=self.cliente_b,
            repartidor=self.other_driver,
            estado='Asignado',
            prioridad='normal',
            peso_total_kg=3,
        )

        self.route = Ruta.objects.create(
            pedido=self.order_assigned,
            repartidor=self.driver,
            latitud_inicio=4.7100,
            longitud_inicio=-74.0700,
            distancia_km=5.2,
            tiempo_estimado_mins=25,
            capacidad_usada_kg=2,
            estado_ruta='asignada',
        )
        RutaParada.objects.create(
            ruta=self.route,
            pedido=self.order_assigned,
            orden=1,
            latitud=4.7110,
            longitud=-74.0720,
            distancia_desde_anterior_km=1.2,
            tiempo_estimado_desde_anterior_mins=6,
        )

    def _authenticate_driver(self):
        response = self.client.post('/api/login/', {
            'username': 'driver-flow',
            'password': '123',
        }, format='json')
        self.assertEqual(response.status_code, 200)
        token = response.data.get('token')
        self.assertTrue(token)
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')
        return token

    def test_login_returns_token_for_driver(self):
        response = self.client.post('/api/login/', {
            'username': 'driver-flow',
            'password': '123',
        }, format='json')

        self.assertEqual(response.status_code, 200)
        self.assertIn('token', response.data)
        self.assertEqual(response.data['role'], 'driver')
        self.assertEqual(response.data['username'], 'driver-flow')
        self.assertEqual(response.data['repartidor_id'], self.repartidor.id)

    def test_driver_me_requires_token(self):
        response = self.client.get('/api/drivers/me/')
        self.assertEqual(response.status_code, 401)

    def test_driver_me_with_token_returns_profile(self):
        self._authenticate_driver()
        response = self.client.get('/api/drivers/me/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['username'], 'driver-flow')
        self.assertEqual(response.data['user_id'], self.driver.id)
        self.assertEqual(response.data['role'], 'driver')
        self.assertTrue(response.data['disponible'])

    def test_toggle_availability_updates_both_repartidor_and_user(self):
        self._authenticate_driver()
        response = self.client.post('/api/drivers/me/toggle-availability/', {
            'disponible': False,
        }, format='json')

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data['disponible'])
        self.assertFalse(response.data['available'])
        self.assertIn('message', response.data)

        self.repartidor.refresh_from_db()
        self.driver.refresh_from_db()
        self.assertFalse(self.repartidor.disponible)
        self.assertEqual(self.driver.estado, 'No disponible')

    def test_my_orders_only_returns_orders_assigned_to_authenticated_driver(self):
        self._authenticate_driver()
        response = self.client.get('/api/drivers/me/orders/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['total'], 1)
        self.assertEqual(len(response.data['pedidos']), 1)
        self.assertEqual(response.data['pedidos'][0]['id'], self.order_assigned.id)

    def test_my_orders_accepts_estado_filter(self):
        self._authenticate_driver()

        response_positive = self.client.get('/api/drivers/me/orders/?estado=Asignado')
        self.assertEqual(response_positive.status_code, 200)
        self.assertEqual(response_positive.data['total'], 1)

        response_negative = self.client.get('/api/drivers/me/orders/?estado=Entregado')
        self.assertEqual(response_negative.status_code, 200)
        self.assertEqual(response_negative.data['total'], 0)

    def test_my_routes_returns_routes_with_stops(self):
        self._authenticate_driver()
        response = self.client.get('/api/drivers/me/routes/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['total'], 1)
        self.assertEqual(len(response.data['rutas']), 1)
        self.assertEqual(response.data['rutas'][0]['id'], self.route.id)
        self.assertEqual(response.data['rutas'][0]['total_paradas'], 1)

    def test_location_update_persists_coordinates_and_last_location_json(self):
        self._authenticate_driver()
        response = self.client.post('/api/drivers/me/location/', {
            'latitud': 4.75012345,
            'longitud': -74.10123456,
        }, format='json')

        self.assertEqual(response.status_code, 200)

        self.repartidor.refresh_from_db()
        self.assertEqual(float(self.repartidor.latitud_actual), 4.75012345)
        self.assertEqual(float(self.repartidor.longitud_actual), -74.10123456)
        self.assertIn('latitud', self.repartidor.ultima_ubicacion)
        self.assertIn('longitud', self.repartidor.ultima_ubicacion)
        self.assertIn('timestamp', self.repartidor.ultima_ubicacion)

    def test_order_start_changes_order_state_to_en_ruta(self):
        self._authenticate_driver()
        response = self.client.post(f'/api/drivers/me/orders/{self.order_assigned.id}/start/', {}, format='json')

        self.assertEqual(response.status_code, 200)
        self.order_assigned.refresh_from_db()
        self.assertEqual(self.order_assigned.estado, 'En ruta')

    def test_order_deliver_changes_order_state_to_entregado(self):
        self._authenticate_driver()
        self.order_assigned.estado = 'En ruta'
        self.order_assigned.save(update_fields=['estado'])

        response = self.client.post(f'/api/drivers/me/orders/{self.order_assigned.id}/deliver/', {
            'comentarios': 'Entregado sin novedad',
        }, format='json')

        self.assertEqual(response.status_code, 200)
        self.order_assigned.refresh_from_db()
        self.assertEqual(self.order_assigned.estado, 'Entregado')
        self.assertIsNotNone(self.order_assigned.fecha_entrega)

    def test_driver_cannot_start_order_of_another_driver(self):
        self._authenticate_driver()
        response = self.client.post(f'/api/drivers/me/orders/{self.order_other_driver.id}/start/', {}, format='json')

        self.assertEqual(response.status_code, 404)

    def test_order_complete_changes_order_state_to_entregado(self):
        self._authenticate_driver()
        response = self.client.post(f'/api/drivers/me/orders/{self.order_assigned.id}/complete/', {}, format='json')

        self.assertEqual(response.status_code, 200)
        self.order_assigned.refresh_from_db()
        self.assertEqual(self.order_assigned.estado, 'Entregado')
        self.assertIsNotNone(self.order_assigned.fecha_entrega)

    def test_driver_me_returns_clear_message_when_no_repartidor_profile(self):
        user_no_profile = CustomUser.objects.create_user(
            username='driver-no-profile',
            password='123',
            role='driver',
            nombre='Driver Sin Perfil',
        )
        token_response = self.client.post('/api/login/', {
            'username': 'driver-no-profile',
            'password': '123',
        }, format='json')
        self.assertEqual(token_response.status_code, 200)
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {token_response.data['token']}")

        response = self.client.get('/api/drivers/me/')
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data['error'], 'Este usuario no tiene perfil de repartidor asociado.')
